"""
Repository Scan Reader / Excel Exporter
----------------------------------------
Reads scan_XXXX.json files produced by the scanner and lets you browse them
interactively or export them to a clean, human-readable Excel workbook.

Matches the ACTUAL schema written by scan_repository():

{
    "scan_metadata": {
        "scan_time": str,
        "root_folder": str,
        "scan_mode": "Full Sweep" | "Extension Query",
        "report_type": "Summary" | "Detailed"
    },
    "repository_statistics": {
        "folders": int,
        "files": int,
        "file_types": {ext: count, ...}
    },
    "performance": {
        "duration_seconds": float,
        "files_per_second": float,
        "items_processed": int
    },

    # Only present when report_type == "Detailed":
    "analysis": {...placeholders...},
    "extensions": {
        ext: [
            {
                "name": str, "path": str, "extension": str,
                "size_bytes": int, "modified": iso-str,
                "lines": int | None, "hidden": bool
            },
            ...
        ],
        ...
    }
}

There is no top-level "version" key and no top-level "files" array in the
real output -- this script no longer assumes either.
"""

from pathlib import Path
import json
import sys

import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SCAN_DIR = Path("scans")
EXPORT_DIR = Path("exports")


# ---------------------------------------------------------------------------
# Loading
# ---------------------------------------------------------------------------

def load_reports():
    reports = sorted(
        SCAN_DIR.glob("scan_*.json"), key=lambda x: x.stat().st_mtime, reverse=True
    )

    if not reports:
        print("No scan reports found in ./scans")
        return None

    print("\nAvailable Reports\n")
    for i, report in enumerate(reports, start=1):
        print(f"{i}. {report.name}")

    while True:
        try:
            choice = int(input("\nChoose report: "))
            if 1 <= choice <= len(reports):
                return reports[choice - 1]
        except ValueError:
            pass
        print("Invalid selection.")


def load_json(path):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def get_next_export_number(export_dir: Path):
    """Find the next sequential export number, mirroring the scanner's
    scan_0001.json numbering convention (report_001.xlsx, report_002.xlsx, ...)."""
    export_dir.mkdir(parents=True, exist_ok=True)

    existing = sorted(export_dir.glob("report_*.xlsx"))

    numbers = []
    for file in existing:
        try:
            numbers.append(int(file.stem.split("_")[1]))
        except Exception:
            pass

    return max(numbers, default=0) + 1


def is_detailed(data):
    """A report has file-level data only if it carries an 'extensions' dict."""
    return isinstance(data.get("extensions"), dict) and len(data["extensions"]) > 0


def flatten_files(data):
    """Flatten data['extensions'] into a single list of file-info dicts."""
    files = []
    for group in data.get("extensions", {}).values():
        files.extend(group)
    return files


# ---------------------------------------------------------------------------
# Views
# ---------------------------------------------------------------------------

def repository_summary(data):
    meta = data.get("scan_metadata", {})
    stats = data.get("repository_statistics", {})
    perf = data.get("performance", {})

    print("\nRepository Summary")
    print("-" * 40)
    print(f"Root Folder : {meta.get('root_folder', 'Unknown')}")
    print(f"Scan Time   : {meta.get('scan_time', 'Unknown')}")
    print(f"Scan Mode   : {meta.get('scan_mode', 'Unknown')}")
    print(f"Report Type : {meta.get('report_type', 'Unknown')}")
    print(f"Folders     : {stats.get('folders', 0)}")
    print(f"Files       : {stats.get('files', 0)}")
    print(f"Extensions  : {len(stats.get('file_types', {}))}")

    if perf:
        print(f"Duration    : {perf.get('duration_seconds', 0):.2f} sec")
        print(f"Files/sec   : {perf.get('files_per_second', 0)}")
        print(f"Items Seen  : {perf.get('items_processed', 0)}")

    return meta, stats, perf


def browse_extensions(data):
    if not is_detailed(data):
        print("\nThis report is Summary only -- no per-file data available.")
        return

    extensions = data["extensions"]
    ext_list = sorted(extensions.keys())

    print()
    for i, ext in enumerate(ext_list, start=1):
        print(f"{i}. {ext:<12} ({len(extensions[ext])})")

    try:
        choice = int(input("\nSelect Extension: "))
        if 1 <= choice <= len(ext_list):
            extension = ext_list[choice - 1]
        else:
            print("Invalid selection.")
            return
    except Exception:
        print("Invalid input.")
        return

    print("\n" + "=" * 60)
    print(f"Files with {extension}")
    print("=" * 60)

    for file in extensions[extension]:
        print(f"\nName      : {file['name']}")
        print(f"Path      : {file['path']}")
        print(f"Lines     : {file['lines']}")
        print(f"Size      : {file['size_bytes']:,} bytes")
        print(f"Modified  : {file['modified']}")
        print(f"Hidden    : {file['hidden']}")
        print("-" * 60)


def search_file(data):
    if not is_detailed(data):
        print("Detailed report required.")
        return

    query = input("\nSearch filename: ").lower()

    found = False
    for file in flatten_files(data):
        if (
            query in file["name"].lower()
            or query in file["path"].lower()
            or query == file["extension"]
        ):
            found = True
            print("\n--------------------------")
            print(f"Name     : {file['name']}")
            print(f"Extension: {file['extension']}")
            print(f"Path     : {file['path']}")
            print(f"Lines    : {file['lines']}")
            print(f"Size     : {file['size_bytes']:,} bytes")
            print(f"Modified : {file['modified']}")

    if not found:
        print("No matching files.")


def largest_files(data):
    if not is_detailed(data):
        print("Detailed report required.")
        return

    files = sorted(flatten_files(data), key=lambda x: x["size_bytes"], reverse=True)

    print("\nLargest Files\n")
    for rank, file in enumerate(files[:10], start=1):
        print(f"{rank}. {file['name']}")
        print(f"   Size     : {file['size_bytes']:,} bytes")
        print(f"   Extension: {file['extension']}")
        print(f"   Path     : {file['path']}")
        print("-" * 60)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def _style_header(ws, n_cols):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"


def _autofit(ws, df):
    for i, col in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()]
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 60)


def export_to_excel(meta, stats, perf, data, output_file=None):
    if output_file is None:
        export_number = get_next_export_number(EXPORT_DIR)
        output_file = EXPORT_DIR / f"report_{export_number:03}.xlsx"
    else:
        output_file = Path(output_file)
        output_file.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:

        # --- Summary sheet -------------------------------------------------
        summary_rows = [
            ("Root Folder", meta.get("root_folder", "Unknown")),
            ("Scan Time", meta.get("scan_time", "Unknown")),
            ("Scan Mode", meta.get("scan_mode", "Unknown")),
            ("Report Type", meta.get("report_type", "Unknown")),
            ("Folders", stats.get("folders", 0)),
            ("Files", stats.get("files", 0)),
            ("Distinct Extensions", len(stats.get("file_types", {}))),
            ("Scan Duration (sec)", round(perf.get("duration_seconds", 0), 2)),
            ("Files / sec", perf.get("files_per_second", 0)),
            ("Items Processed", perf.get("items_processed", 0)),
        ]
        summary_df = pd.DataFrame(summary_rows, columns=["Field", "Value"])
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        _style_header(ws, 2)
        _autofit(ws, summary_df)

        # --- File Types sheet ------------------------------------------------
        file_types = stats.get("file_types", {})
        if file_types:
            ft_df = pd.DataFrame(
                sorted(file_types.items(), key=lambda x: x[1], reverse=True),
                columns=["Extension", "Count"],
            )
            ft_df.to_excel(writer, sheet_name="File Types", index=False)
            ws = writer.sheets["File Types"]
            _style_header(ws, 2)
            _autofit(ws, ft_df)

        # --- Files sheet (only if Detailed report) ---------------------------
        if is_detailed(data):
            file_records = []
            for f in flatten_files(data):
                file_records.append(
                    {
                        "Name": f["name"],
                        "Path": f["path"],
                        "Extension": f["extension"],
                        "Size (bytes)": f["size_bytes"],
                        "Size (KB)": round(f["size_bytes"] / 1024, 2),
                        "Lines": f["lines"] if f["lines"] is not None else "",
                        "Modified": f["modified"],
                        "Hidden": f["hidden"],
                    }
                )

            files_df = pd.DataFrame(file_records).sort_values("Path").reset_index(drop=True)
            files_df.to_excel(writer, sheet_name="Files", index=False)
            ws = writer.sheets["Files"]
            _style_header(ws, len(files_df.columns))
            _autofit(ws, files_df)
            ws.auto_filter.ref = ws.dimensions
        else:
            note_df = pd.DataFrame(
                [["This is a Summary report -- no per-file data was captured. "
                  "Re-run the scanner with Report Type = Detailed to get a Files sheet."]],
                columns=["Note"],
            )
            note_df.to_excel(writer, sheet_name="Files", index=False)
            ws = writer.sheets["Files"]
            _style_header(ws, 1)
            _autofit(ws, note_df)

    print(f"\nExcel report exported to: {output_file.resolve()}")


# ---------------------------------------------------------------------------
# Menu
# ---------------------------------------------------------------------------

def menu(data):
    meta, stats, perf = repository_summary(data)

    while True:
        print("\nRepository Reader")
        print("---------------------------")
        print("1. Repository Summary")
        print("2. Browse Extensions")
        print("3. Search File")
        print("4. Largest Files")
        print("5. Export to Excel")
        print("6. Exit")

        choice = input("\nChoice: ").strip()

        if choice == "1":
            repository_summary(data)
        elif choice == "2":
            browse_extensions(data)
        elif choice == "3":
            search_file(data)
        elif choice == "4":
            largest_files(data)
        elif choice == "5":
            export_to_excel(meta, stats, perf, data)
        elif choice == "6":
            break
        else:
            print("Invalid choice.")


def main():
    report = load_reports()
    if report is None:
        sys.exit(0)
    data = load_json(report)
    menu(data)


if __name__ == "__main__":
    main()
