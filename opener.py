"""
opener.py
---------
Third stage of the Iron2Tech repo scanner pipeline:

    scanner.py  -> scans a folder, writes scans/scan_XXXX.json
    reader.py   -> reads a scan_XXXX.json, exports exports/report_XXX.xlsx
    opener.py   -> reads EITHER of the above, lets you search for a file
                   by name and open it (double-click equivalent)

How it works:
    1. On startup, it looks in ./scans and ./exports for scan_*.json and
       report_*.xlsx files and lets you pick one (or paste a path to a
       different file). That file is loaded ONCE.
    2. From then on you just type a keyword (e.g. "IRON2TECH") and it
       searches every indexed filename/path for that keyword.
         - No matches  -> prints a plain "couldn't find that" message.
         - One match   -> opens it immediately.
         - Several     -> lists them and asks which one you meant.
    3. The loaded index lives only in memory for this run. Nothing is
       cached to disk, so closing the program clears it -- next run
       starts fresh and asks you to pick a source file again.

Scope note: the scanner currently only indexes individual FILES (folders
are only ever a total count in repository_statistics.folders), so this
version searches files only. Folder search can be added once the scanner
is extended to list individual folder paths.

Important schema detail: paths stored in the scan/report are RELATIVE to
scan_metadata.root_folder (JSON) / the "Root Folder" row on the Summary
sheet (Excel). This script joins root_folder + relative path before
trying to open anything.
"""

from pathlib import Path
import json
import os
import platform
import subprocess

SCAN_DIR = Path("scans")
EXPORT_DIR = Path("exports")


# ---------------------------------------------------------------------------
# Picking a source file (scan_XXXX.json or report_XXX.xlsx)
# ---------------------------------------------------------------------------

def choose_source():
    scans = sorted(SCAN_DIR.glob("scan_*.json"), reverse=True) if SCAN_DIR.exists() else []
    reports = sorted(EXPORT_DIR.glob("report_*.xlsx"), reverse=True) if EXPORT_DIR.exists() else []
    combined = [("scan", p) for p in scans] + [("report", p) for p in reports]

    if combined:
        print("\nAvailable scan/report files:")
        for i, (kind, p) in enumerate(combined, start=1):
            print(f"  {i}. [{kind}] {p.name}")
        prompt = "\nChoose a number, or paste a path to a different .json/.xlsx file: "
    else:
        print("\nNo files found in ./scans or ./exports.")
        prompt = "Enter a path to a .json or .xlsx file: "

    choice = input(prompt).strip()
    if not choice:
        return None

    if choice.isdigit() and combined:
        idx = int(choice) - 1
        if 0 <= idx < len(combined):
            return combined[idx][1]
        print("That number isn't in the list.")
        return None

    return Path(choice)


# ---------------------------------------------------------------------------
# Loading -- returns (records, root_folder, note)
# records: list of dicts with name / rel_path / abs_path / extension / size_bytes / modified
# note: a human-readable reason there's nothing to search, or None
# ---------------------------------------------------------------------------

def load_index(source_path):
    p = Path(source_path)
    if not p.exists():
        raise FileNotFoundError(f"That file doesn't exist: {p}")

    if p.suffix.lower() == ".json":
        return _load_from_scan_json(p)
    elif p.suffix.lower() in (".xlsx", ".xlsm"):
        return _load_from_excel_report(p)
    else:
        raise ValueError(f"Unsupported file type '{p.suffix}' -- expected .json or .xlsx")


def _load_from_scan_json(p):
    with open(p, "r", encoding="utf-8") as f:
        data = json.load(f)

    meta = data.get("scan_metadata", {})
    root_folder = meta.get("root_folder", "")
    extensions = data.get("extensions")

    if not extensions:
        note = ("This scan has no per-file index (it's a Summary report). "
                "Re-run the scanner with Report Type = Detailed to search individual files.")
        return [], root_folder, note

    records = []
    for group in extensions.values():
        for f in group:
            rel_path = f.get("path", "")
            records.append({
                "name": f.get("name", Path(rel_path).name),
                "rel_path": rel_path,
                "abs_path": str(Path(root_folder) / rel_path) if root_folder else rel_path,
                "extension": f.get("extension", ""),
                "size_bytes": f.get("size_bytes"),
                "modified": f.get("modified"),
            })
    return records, root_folder, None


def _load_from_excel_report(p):
    from openpyxl import load_workbook

    wb = load_workbook(p, read_only=True, data_only=True)

    root_folder = ""
    if "Summary" in wb.sheetnames:
        for row in wb["Summary"].iter_rows(values_only=True):
            if row and row[0] == "Root Folder":
                root_folder = row[1] or ""
                break

    if "Files" not in wb.sheetnames:
        return [], root_folder, "This workbook has no 'Files' sheet."

    rows = wb["Files"].iter_rows(values_only=True)
    header = [str(h).strip() if h else "" for h in next(rows)]

    if header == ["Note"]:
        # report_reader.py writes this single-column sheet for Summary-only reports
        return [], root_folder, ("This report has no per-file data (it's a Summary report). "
                                  "Re-run the scanner with Report Type = Detailed and re-export.")

    col_idx = {h: i for i, h in enumerate(header)}
    if "Path" not in col_idx:
        raise ValueError(f"'Files' sheet doesn't have a 'Path' column. Found columns: {header}")

    records = []
    for row in rows:
        rel_path = row[col_idx["Path"]]
        if rel_path is None:
            continue
        name = row[col_idx["Name"]] if "Name" in col_idx else Path(rel_path).name
        records.append({
            "name": name,
            "rel_path": rel_path,
            "abs_path": str(Path(root_folder) / rel_path) if root_folder else rel_path,
            "extension": row[col_idx["Extension"]] if "Extension" in col_idx else "",
            "size_bytes": row[col_idx["Size (bytes)"]] if "Size (bytes)" in col_idx else None,
            "modified": row[col_idx["Modified"]] if "Modified" in col_idx else None,
        })
    return records, root_folder, None


# ---------------------------------------------------------------------------
# Search + open
# ---------------------------------------------------------------------------

def search(records, keyword):
    keyword = keyword.lower()
    return [
        r for r in records
        if keyword in r["name"].lower()
        or keyword in r["rel_path"].lower()
        or keyword in r["abs_path"].lower()
    ]


def open_record(record):
    p = Path(record["abs_path"])
    if not p.exists():
        print(f"Found '{record['name']}' in the index, but it no longer exists on disk:")
        print(f"  {p}")
        print("  (It may have been moved, renamed, or deleted since the last scan.)")
        return

    try:
        system = platform.system()
        if system == "Windows":
            os.startfile(str(p))  # noqa: only exists on Windows
        elif system == "Darwin":
            subprocess.run(["open", str(p)], check=True)
        else:
            subprocess.run(["xdg-open", str(p)], check=True)
        print(f"Opened: {p}")
    except Exception as e:
        print(f"Found '{record['name']}' but couldn't open it: {e}")


# ---------------------------------------------------------------------------
# Main loop
# ---------------------------------------------------------------------------

def main():
    try:
        source = choose_source()
    except Exception as e:
        print(f"Something went wrong choosing a file: {e}")
        return

    if source is None:
        print("No file selected. Exiting.")
        return

    try:
        records, root_folder, note = load_index(source)
    except Exception as e:
        print(f"Couldn't load that file: {e}")
        return

    if note:
        print(f"\n{note}")
        return

    print(f"\nLoaded {len(records)} indexed files.")
    print(f"Root folder: {root_folder or '(not recorded)'}")
    print("Type a filename or partial name to search. Type 'quit' to exit.\n")

    while True:
        keyword = input("Search: ").strip()

        if keyword.lower() in ("quit", "exit", "q"):
            print("Goodbye.")
            break
        if not keyword:
            continue

        matches = search(records, keyword)

        if not matches:
            print(f"Could not find any file matching '{keyword}'.\n")
            continue

        if len(matches) == 1:
            open_record(matches[0])
            print()
            continue

        print(f"\nFound {len(matches)} matches:")
        for i, r in enumerate(matches, start=1):
            print(f"  {i}. {r['rel_path']}")

        choice = input(f"Which one? (1-{len(matches)}, or blank to cancel): ").strip()
        if not choice:
            print("Cancelled.\n")
            continue

        if choice.isdigit() and 1 <= int(choice) <= len(matches):
            open_record(matches[int(choice) - 1])
        else:
            print("Invalid selection.")
        print()


if __name__ == "__main__":
    main()
